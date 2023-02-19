from rest_framework import generics
from API.models import Page, Review, Check, Report, Subscription, User, CheckReport
from API.serializers.page import PageSerializer
from rest_framework.permissions import AllowAny, IsAuthenticated
from API.permissions import IsAdmin
from django.http import JsonResponse
from django.core.paginator import Paginator
from API.funcs import getData
from django.core.exceptions import ValidationError
from django.utils import timezone
from datetime import timedelta
from os import remove
from re import search
from ping3 import ping as ping3
from urllib.parse import urlparse
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from project.settings import host

errors = {'500': {'error_description': 'Ошибка сервера',
                  'reasons': [
                      'Перезагрузка сервера',
                      'Ошибка в коде сервера',
                      'Ошибка в файле .htaccess']},
          '520': {'error_description': 'Неизвестная ошибка',
                  'reasons': ['Сайт обрывает соединение']},
          '522': {'error_description': 'Время ожидания подключения к исходному серверу истекло',
                  'reasons': ['Сервер не работает']},
          '524': {'error_description': 'Время ожидания ответа сервера истекло',
                  'reasons': ['Сервер не отправляет своевременный HTTP-ответ']},
          '404': {'error_description': 'Странице не найдена',
                  'reasons': ['Страница была удалена']},
          '400': {'error_description': 'Ошибка клиента',
                  'reasons': ['Неверная ссылка']},
          '200': {'error_description': 'Медленная загрузка',
                  'reasons': ['Большой трафик', 'DDOS', 'т.д.']}}


class PageListView(generics.ListAPIView):
    serializer_class = PageSerializer
    permission_classes = [IsAuthenticated]
    queryset = Page.objects.all()


class PageCreate(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PageSerializer

    def post(self, request, *args, **kwargs):
        try:
            data = getData(request)
            user = request.user
            try:
                if bool(search('[а-яА-Я]', data['url'])):
                    if data['url'][-1] != '/':
                        url = data['url'] + '/'
                    else:
                        url = data['url']
                else:
                    url = requests.get(data['url']).url
            except Exception:
                url = data['url']
            page = list(Page.objects.filter(url=url).exclude(is_moderated=False))
            if page:
                return JsonResponse({'errors': {'url': ['Сайт с таким доменом уже есть в базе']}})
            page = Page(name=data['name'],
                        url=url,
                        description=data['description'],
                        added_by_user_id=user.id)
            try:
                page.full_clean()
            except ValidationError as ex:
                return JsonResponse({'errors': dict(ex)})
            page.save()
            return JsonResponse({'success': True})
        except Exception:
            return JsonResponse({'success': False}, status=400)


class PageRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdmin]
    serializer_class = PageSerializer
    queryset = Page.objects.all()


class GetPopularPages(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = PageSerializer

    def get(self, request, *args, **kwargs):
        try:
            reviews = Review.objects.raw('SELECT pages.id, pages.name, pages.url, COUNT(CASE WHEN revs.is_published=true THEN 1 ELSE NULL END) AS "total" '
                                         'FROM "API_page" AS pages '
                                         'LEFT JOIN "API_review" AS revs ON revs.page_id=pages.id '
                                         'WHERE pages.is_checking = true '
                                         'GROUP BY pages.id '
                                         'ORDER BY total DESC, pages.id DESC '
                                         'LIMIT 3 ')
            pageIds = [i.id for i in reviews]
            checks = list(Page.objects.filter(id__in=pageIds)
                          .select_related('checks')
                          .values('id', 'checks__check_status')
                          .order_by('-id', '-checks__id')
                          .distinct('id'))
            temp = {}
            for i in checks:
                temp[i['id']] = i['checks__check_status']
            res = []
            for i in reviews:
                res.append({'id': i.id,
                            'name': i.name,
                            'url': i.url,
                            'check_status': temp[i.id]})
            return JsonResponse(res, safe=False)
        except Exception:
            return JsonResponse({'success': False}, status=500)


class GetSiteStats(generics.GenericAPIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        try:
            pages = Page.objects.only('id').count()
            reports = Report.objects.only('id').count()
            reviews = Review.objects.only('id').count()
            failures = Check.objects.only('id').exclude(response_status_code=200).count()
            result = {'total_pages': pages, 'total_reports': reports, 'total_reviews': reviews,
                      'detected_failures': failures}
            return JsonResponse(result)
        except Exception:
            return JsonResponse({'success': False}, status=500)


class GetCheckingPages(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = PageSerializer

    def post(self, request, *args, **kwargs):
        try:
            data = getData(request)
            pages = Page.objects.raw('SELECT DISTINCT ON (pages.id) pages.id, pages.name, pages.url, '
                                     'checks.response_time, checks.checked_at, checks.check_status, '
                                     'ROUND(AVG(reviews.mark), 2) as rating '
                                     'FROM "API_page" AS pages '
                                     'LEFT JOIN "API_check" AS checks ON checks.page_id=pages.id '
                                     'LEFT JOIN "API_review" AS reviews ON (pages.id=reviews.page_id AND reviews.is_published=true) '
                                     'WHERE pages.is_checking = true '
                                     'GROUP BY pages.id, checks.response_time, checks.checked_at, checks.check_status '
                                     'ORDER BY pages.id DESC, checks.checked_at DESC ')
            result = []
            for i in pages:
                if data['search'] == '' or (data['search'].lower() in i.name.lower()) or \
                        (data['search'].lower() in i.description.lower() or \
                         (data['search'].lower() in i.url.lower())):
                    result.append({'id': i.id,
                                   'name': i.name,
                                   'url': i.url,
                                   'last_check_time': i.checked_at,
                                   'last_check_timeout': i.response_time,
                                   'check_status': i.check_status,
                                   'rating': i.rating})
            result = Paginator(result, 10)
            return JsonResponse({'num_pages': result.num_pages, 'pages': list(result.page(data['page_number']))})
        except Exception:
            return JsonResponse({'success': False}, status=500)


class GetAccountData(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            user = request.user
            subs = Subscription.objects.raw(
                'SELECT DISTINCT ON (pages.id) pages.id, pages.name, checks.check_status '
                'FROM "API_subscription" as subs '
                'LEFT JOIN "API_check" as checks ON checks.page_id=subs.page_id '
                'JOIN "API_page" as pages ON subs.page_id=pages.id '
                f'WHERE user_id={user.id} '
                'ORDER BY pages.id DESC, checks.page_id DESC, checks.id DESC')
            result = []
            for i in subs:
                result.append({'id': i.id,
                               'name': i.name,
                               'check_status': i.check_status})
            return JsonResponse(result, safe=False)
        except Exception:
            return JsonResponse({'success': False}, status=500)


class PageChecks(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = PageSerializer

    def get(self, request, id, *args, **kwargs):
        try:
            time = timezone.now()
            checks = list(Check.objects.filter(page_id=id, checked_at__range=(time - timedelta(days=3), time))
                          .values('checked_at', 'response_time', 'check_status')
                          .order_by('id'))
            return JsonResponse(checks, safe=False)
        except Exception:
            return JsonResponse({'detail': 'Something went wrong'})


class PageReviews(generics.GenericAPIView):
    serializer_class = PageSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        elif self.request.method == 'POST':
            return [IsAuthenticated()]
        else:
            return [IsAdmin()]

    def get(self, request, id, *args, **kwargs):
        try:
            checks = list(Review.objects.filter(page_id=id)
                          .select_related('added_by_user')
                          .values('id', 'added_at', 'mark', 'message', 'added_by_user__username', 'added_by_user')
                          .filter(is_published=True))
            return JsonResponse(checks, safe=False)
        except Exception:
            return JsonResponse({'detail': 'Something went wrong'})

    def post(self, request, id, *args, **kwargs):
        try:
            user = request.user
            data = getData(request)
            review = Review.objects.create(added_by_user_id=user.id, page_id=id, message=data['message'],
                                           mark=data['mark'])
            review.save()
            return JsonResponse({'success': True})
        except Exception:
            return JsonResponse({'success': False}, status=400)


class PageReports(generics.GenericAPIView):
    serializer_class = PageSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        elif self.request.method == 'POST':
            return [IsAuthenticated()]
        else:
            return [IsAdmin()]

    def get(self, request, id, *args, **kwargs):
        try:
            checks = list(Report.objects.filter(page_id=id)
                          .select_related('added_by_user')
                          .values('id', 'added_at', 'message', 'added_by_user__username', 'added_by_user')
                          .filter(is_published=True))
            return JsonResponse(checks, safe=False)
        except Exception:
            return JsonResponse({'detail': 'Something went wrong'}, status=400)

    def post(self, request, id, *args, **kwargs):
        try:
            user = request.user
            data = getData(request)
            report = Report.objects.create(added_by_user_id=user.id, page_id=id, message=data['message'],
                                           added_at=data['added_at'])
            report.save()
            return JsonResponse({'success': True})
        except Exception:
            return JsonResponse({'success': False}, status=400)


class GetPageData(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = PageSerializer

    def get(self, request, id, *args, **kwargs):
        try:
            page = list(Page.objects.filter(id=id).values('name', 'url', 'description'))
            if page:
                page = page[0]
            return JsonResponse(page, safe=False)
        except Exception:
            return JsonResponse({'detail': 'Something went wrong'}, status=400)


class Subscriptions(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = Subscription

    def get(self, request, id, *args, **kwargs):
        try:
            user = request.user
            check_sub = list(Subscription.objects.filter(page_id=id, user_id=user.id))
            if check_sub:
                return JsonResponse({'subscribed': True})
            else:
                return JsonResponse({'subscribed': False})
        except Exception:
            return JsonResponse({'detail': 'Exception'}, status=400)

    def post(self, request, id, *args, **kwargs):
        try:
            user = request.user
            check_sub = list(Subscription.objects.filter(page_id=id, user_id=user.id))
            if not check_sub:
                sub = Subscription.objects.create(page_id=id, user_id=user.id)
                sub.save()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False})
        except Exception:
            return JsonResponse({'detail': 'Exception'}, status=400)

    def delete(self, request, id, *args, **kwargs):
        try:
            user = request.user
            check_sub = Subscription.objects.filter(page_id=id, user_id=user.id)
            if list(check_sub):
                check_sub.delete()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False})
        except Exception:
            return JsonResponse({'detail': 'Exception'}, status=500)


class Events(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PageSerializer

    def get(self, request):
        try:
            user = request.user
            result = {'working': 0,
                      'lazy_loading': 0,
                      'not_working': 0}
            subPages = Subscription.objects.filter(user_id=user.id).values('page_id')
            pageIds = [i['page_id'] for i in subPages]
            time = timezone.now()
            reports = Report.objects.filter(page_id__in=pageIds,
                                            is_published=True,
                                            added_at__range=(time - timedelta(days=3), time)) \
                .select_related('page', 'added_by_user') \
                .values('page', 'page__name', 'added_at', 'message', 'added_by_user', 'added_by_user__username')
            reviews = Review.objects.filter(page_id__in=pageIds,
                                            is_published=True,
                                            added_at__range=(time - timedelta(days=3), time)) \
                .select_related('page', 'added_by_user') \
                .values('page', 'page__name', 'added_at', 'mark', 'message', 'added_by_user', 'added_by_user__username')

            checks = Check.objects.filter(page_id__in=pageIds,
                                          checked_at__range=(time - timedelta(days=3), time)) \
                .select_related('page') \
                .exclude(check_status='2') \
                .values('page_id', 'page__name', 'check_status', 'checked_at', 'response_status_code', 'response_time',
                        'check_status') \
                .order_by('page_id', '-id')

            pages = Check.objects.filter(page_id__in=pageIds) \
                .values('check_status') \
                .distinct('page_id') \
                .order_by('page_id', '-id')
            for i in pages:
                if i['check_status'] == '2':
                    result['working'] += 1
                if i['check_status'] == '1':
                    result['lazy_loading'] += 1
                elif i['check_status'] == '0' or i['check_status'] == '3':
                    result['not_working'] += 1

            result['events'] = []

            for i in checks:
                if i['response_status_code'] == '200' and i['check_status'] == '2':
                    continue
                elif i['check_status'] == '1':
                    result['events'].append({'type': 'lazy_loading',
                                             'page': {'id': i['page_id'],
                                                      'name': i['page__name']},
                                             'detail': {'time': i['response_time'],
                                                        'reasons': [
                                                            'Большой трафик',
                                                            'DDOS',
                                                            'т.д.']},
                                             'message_datetime': i['checked_at']})
                else:
                    try:
                        error = errors[i.response_status_code]
                    except Exception:
                        error = {'error_description': 'Неизвестная ошибка',
                                 'reasons': ['Не известны']}
                    result['events'].append({'type': 'failure',
                                             'page': {'id': i['page_id'],
                                                      'name': i['page__name']},
                                             'detail': error,
                                             'message_datetime': i['checked_at']})
            for i in reports:
                result['events'].append({'type': 'report',
                                         'page': {'id': i['page'],
                                                  'name': i['page__name']},
                                         'detail': {
                                             'message': i['message'],
                                             'user': {
                                                 'id': i['added_by_user'],
                                                 'username': i['added_by_user__username']
                                             }
                                         },
                                         'message_datetime': i['added_at']
                                         })

            for i in reviews:
                result['events'].append({'type': 'review',
                                         'page': {'id': i['page'],
                                                  'name': i['page__name']},
                                         'detail': {
                                             'mark': i['mark'],
                                             'message': i['message'],
                                             'user': {
                                                 'id': i['added_by_user'],
                                                 'username': i['added_by_user__username']
                                             }
                                         },
                                         'message_datetime': i['added_at']
                                         })
            return JsonResponse(result, safe=False)

        except Exception:
            return JsonResponse({'success': False}, status=500)


class DeepCheck(generics.GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = CheckReport

    def post(self, request, level):
        data = getData(request)

        def generate_report(checkreport, checks=None, reports=None):
            try:
                greenFill = PatternFill(start_color='c6efce',
                                        end_color='c6efce',
                                        fill_type='solid')
                greenFont = Font(color='006100')
                yellowFill = PatternFill(start_color='ffeb9c',
                                         end_color='ffeb9c',
                                         fill_type='solid')
                yellowFont = Font(color='9c573e')
                redFill = PatternFill(start_color='ffc7ce',
                                      end_color='ffc7ce',
                                      fill_type='solid')
                redFont = Font(color='ad0006')
                right = Alignment(horizontal='right', vertical='bottom')
                if checkreport.page_id is not None:
                    report = openpyxl.load_workbook('static/templates/report with page.xlsx')
                else:
                    report = openpyxl.load_workbook('static/templates/report without page.xlsx')
                report.active = report['Главное']
                general = report.active
                for cell in general["B:B"]:
                    cell.alignment = right
                general['B2'].hyperlink = checkreport.requested_url
                general['B2'].value = checkreport.requested_url
                general['B3'].value = checkreport.ping
                general['B4'].value = int(checkreport.response_status_code)
                general['B5'].value = checkreport.response_time
                if checkreport.response_status_code.startswith('2') or checkreport.response_status_code.startswith('3'):
                    if checkreport.response_time >= 1000:
                        general['B6'].value = 'Работает медленно'
                        general['B6'].fill = yellowFill
                        general['B6'].font = yellowFont
                    else:
                        general['B6'].value = 'Работает'
                        general['B6'].fill = greenFill
                        general['B6'].font = greenFont
                else:
                    general['B6'].value = 'Не работает'
                    general['B6'].fill = redFill
                    general['B6'].font = redFont
                general['B9'].value = (
                    checkreport.first_content_loading_time if checkreport.first_content_loading_time is not None else 'Нет данных')
                general['B10'].value = (
                    checkreport.first_meaningful_content_loading_time if checkreport.first_meaningful_content_loading_time is not None else 'Нет данных')
                general['B11'].value = (
                    checkreport.largest_content_loading_time if checkreport.largest_content_loading_time is not None else 'Нет данных')
                general['B12'].value = (
                    checkreport.speed_index if checkreport.speed_index is not None else 'Нет данных')
                general['B13'].value = (
                    checkreport.full_page_loading_time if checkreport.full_page_loading_time is not None else 'Нет данных')
                general['B14'].value = (checkreport.score if checkreport.score is not None else 'Нет данных')
                if checkreport.score is None:
                    general['B14'].value = 'Нет данных'
                elif checkreport.score >= 90:
                    general['B14'].font = greenFont
                    general['B14'].fill = greenFill
                elif checkreport.score >= 50:
                    general['B14'].font = yellowFont
                    general['B14'].fill = yellowFill
                else:
                    general['B14'].font = redFont
                    general['B14'].fill = redFill
                if checkreport.page_id is not None:
                    if checks:
                        start = 2
                        report.active = report['Результаты проверок']
                        general = report.active
                        for i in checks:
                            general[f'A{str(start)}'].value = (i['checked_at'].replace(tzinfo=None))
                            general[f'A{str(start)}'].number_format = 'yyyy-mm-dd hh:mm:ss'
                            general[f'B{str(start)}'].value = int(i['response_status_code'])
                            general[f'C{str(start)}'].value = i['response_time']
                            if i['response_status_code'].startswith('2') or i['response_status_code'].startswith('3'):
                                general[f'B{str(start)}'].font = greenFont
                                general[f'B{str(start)}'].fill = greenFill
                            else:
                                general[f'B{str(start)}'].font = redFont
                                general[f'B{str(start)}'].fill = redFill
                            start += 1
                    if reports:
                        start = 2
                        report.active = report['Сообщения о сбоях']
                        general = report.active
                        for i in reports:
                            general[f'A{str(start)}'].value = (i['added_at'].replace(tzinfo=None))
                            general[f'A{str(start)}'].number_format = 'yyyy-mm-dd hh:mm:ss'
                            general[f'B{str(start)}'].value = i['message']
                            start += 1
                report.save(f'temp_files/reports/{check_report.id}.xlsx')
                return True
            except Exception:
                return False

        try:
            if level == 1:
                try:
                    headers = requests.utils.default_headers()
                    headers.update({'User-Agent': 'My User Agent 1.0', })
                    try:
                        if bool(search('[а-яА-Я]', data['url'])):
                            if data['url'][-1] != '/':
                                url = data['url'] + '/'
                            else:
                                url = data['url']
                        else:
                            url = requests.get(data['url']).url
                    except Exception:
                        url = data['url']
                    response_time = 0
                    try:
                        response = requests.get(url, headers=headers)
                        response_code = response.status_code
                        response_time = round(response.elapsed.total_seconds() * 1000)
                    except Exception:
                        response_code = '524'
                    domain = str(urlparse(url).hostname)
                    ping = ping3(domain)
                    if ping == 0:
                        ping = 0
                        response_code = '-'
                    elif ping is None:
                        ping = 0
                    else:
                        ping = round(ping * 1000)
                    check_report = CheckReport(requested_url=url,
                                               ping=ping,
                                               response_status_code=str(response_code),
                                               response_time=response_time)
                    has_data = False
                    page = list(Page.objects.filter(url=url, is_checking=True).values('id'))
                    if page:
                        check_report.page_id = page[0]['id']
                        has_data = True
                    check_report.save()
                    return JsonResponse({'id': check_report.id,
                                         'page_id': check_report.page_id,
                                         'has_data': has_data,
                                         'ping': ping,
                                         'response_status_code': response_code,
                                         'response_time': response_time})
                except Exception as ex:
                    print(ex)
                    return JsonResponse({'success': False}, status=500)
            elif level == 2:
                try:
                    check_report = CheckReport.objects.get(id=int(data['id']))
                    req = f"https://lifegame.su/crasher/proxy.php?url={check_report.requested_url}"
                    GPSI = requests.get(req).json()
                    try:
                        first_content_loading_time = round(GPSI['lighthouseResult']['audits']['first-contentful-paint'][
                            'numericValue'])
                    except Exception:
                        first_content_loading_time = None
                    try:
                        first_meaningful_content_loading_time = \
                            round(GPSI['lighthouseResult']['audits']['first-meaningful-paint']['numericValue'])
                    except Exception:
                        first_meaningful_content_loading_time = None
                    try:
                        largest_content_loading_time = round(GPSI['lighthouseResult']['audits']['largest-contentful-paint'][
                            'numericValue'])
                    except Exception:
                        largest_content_loading_time = None
                    try:
                        speed_index = round(GPSI['lighthouseResult']['audits']['speed-index']['numericValue'])
                    except Exception:
                        speed_index = None
                    try:
                        score = round(GPSI['lighthouseResult']['audits']['speed-index']['score'] * 100)
                    except Exception:
                        score = None
                    try:
                        full_page_loading_time = round(GPSI['lighthouseResult']['timing']['total'])
                    except Exception:
                        full_page_loading_time = None
                    check_report.first_content_loading_time = first_content_loading_time
                    check_report.first_meaningful_content_loading_time = first_meaningful_content_loading_time
                    check_report.largest_content_loading_time = largest_content_loading_time
                    check_report.speed_index = speed_index
                    check_report.score = score
                    check_report.full_page_loading_time = full_page_loading_time
                    check_report.save()
                    return JsonResponse({'first_contentful_paint': first_content_loading_time,
                                         'first_meaningful_paint': first_meaningful_content_loading_time,
                                         'largest_contentful_paint': largest_content_loading_time,
                                         'speed_index': speed_index,
                                         'score': score,
                                         'full_page_loading_time': full_page_loading_time})
                except Exception:
                    return JsonResponse({'success': False}, status=500)
            elif level == 3:
                try:
                    check_report = CheckReport.objects.get(id=data['id'])
                    if check_report.page_id is not None:
                        date_from = data['date_from']
                        date_to = data['date_to']
                        checks = list(Check.objects.filter(checked_at__range=(date_from, date_to),
                                                           page_id=check_report.page_id)
                                      .values('response_time', 'checked_at', 'response_status_code'))
                        reports = list(Report.objects.filter(added_at__range=(date_from, date_to),
                                                             page_id=check_report.page_id,
                                                             is_published=True)
                                       .values('added_at', 'message'))
                        generate_report(check_report, checks, reports)
                    else:
                        generate_report(check_report)
                    file = open(f'temp_files/reports/{check_report.id}.xlsx', 'rb')
                    check_report.report_file.save(f'{check_report.id}.xlsx', file)
                    file.close()
                    remove(f'temp_files/reports/{check_report.id}.xlsx')
                    check_reports = list(CheckReport.objects.filter(requested_url=check_report.requested_url)
                                         .exclude(report_file="")
                                         .values('report_file', 'created_at')
                                         .order_by('-id', 'created_at')[1:4])
                    other_check_reports = []
                    if check_reports:
                        for i in check_reports:
                            url = f"{host}/media/{i['report_file']}"
                            other_check_reports.append({'date': i['created_at'],
                                                        'document_url': url})
                    return JsonResponse({'document_url': f'{host}/media/{check_report.report_file}',
                                         'other_check_reports': other_check_reports})
                except Exception:
                    return JsonResponse({'success': False}, status=500)
        except Exception:
            return JsonResponse({'success': False}, status=400)
